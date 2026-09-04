"""网页版入口：上传表格 -> 选模式 -> 搭建。

两种模式共用这一套界面和进度显示，区别只在于用哪个 builder、哪个浏览器 profile：

  minigame  小游戏        src/builder.py          browser_profile
  drama     短剧商品库     src/drama/builder.py    browser_profile_drama

两个 profile 必须分开：它们是两次独立的登录会话，混用会互相顶掉登录状态。
"""

import threading
import traceback

from flask import Flask, jsonify, render_template, request
from playwright.sync_api import sync_playwright

from src.account import (
    check_advertiser_access,
    clear_profile,
    describe_access,
    is_browser_closed_error,
    profile_has_login,
    wait_for_login,
)
from src.builder import build_campaign_group
from src.config import (
    ACCEPT_LANGUAGE,
    BROWSER_PROFILE_DIR,
    LOCALE,
    UPLOADS_DIR,
)
from src.drama.builder import build_drama_campaign
from src.drama.config import DRAMA_BROWSER_PROFILE_DIR
from src.episode.builder import build_episode_campaign
from src.episode.config import EPISODE_BROWSER_PROFILE_DIR
from src.drama.series_lookup import load_series_map, resolve_series_from_campaign_name
from src.excel_loader import group_by_campaign, load_rows
from src.identity_lookup import IDENTITY_FILE, identity_file_exists, load_identity_map

app = Flask(__name__)

UPLOAD_PATH = UPLOADS_DIR / "latest.xlsx"

# 不要用 5000：macOS 的「隔空播放接收器」（ControlCenter）常驻占着 *:5000，
# Flask 起不来会报 Address already in use，而浏览器可能连到别的残留进程上，
# 看到的是旧页面——排查起来非常费劲。
PORT = 5050

MODES = {
    "minigame": {
        "label": "小游戏",
        "profile": BROWSER_PROFILE_DIR,
    },
    "drama": {
        "label": "短剧商品库",
        "profile": DRAMA_BROWSER_PROFILE_DIR,
    },
    "episode": {
        "label": "短剧端计划",
        # 登录态和商品库共用（实测同一个 Business Center，见 episode/config.py）
        "profile": EPISODE_BROWSER_PROFILE_DIR,
    },
}

state_lock = threading.Lock()
run_state = {
    "status": "idle",  # idle | running | done | error
    "mode": None,
    "total": 0,
    "completed": 0,
    "current": None,
    "results": [],
    "fatal_error": None,
}

# 登录窗口是另一条长时间运行的线程，状态单独放。
# 必须和搭建互斥：Chromium 会锁住 user_data_dir，两边同时开同一个 profile 会打架。
login_state = {
    "status": "idle",  # idle | waiting | done | error
    "mode": None,
    "message": "",
}

# 自动共享素材。和搭建是两码事（不建计划、只在素材库里操作），所以状态单独放；
# 但它们共用同一个浏览器 profile，所以必须互斥 —— 见 _busy_reason。
share_state = {
    "status": "idle",   # idle | running | done | error
    "mode": None,
    "total": 0,
    "completed": 0,
    "current": None,
    "results": [],
    "fatal_error": None,
}


def _reset_state():
    run_state.update(
        status="idle",
        mode=None,
        total=0,
        completed=0,
        current=None,
        results=[],
        fatal_error=None,
    )


def _busy_reason():
    """现在有没有别的事情正占着浏览器 profile。没有就返回 None。"""
    with state_lock:
        if run_state["status"] == "running":
            return "正在搭建，等它跑完再操作账号"
        if login_state["status"] == "waiting":
            return "登录窗口还开着，先在那个窗口里登录完（或者关掉它）"
        if share_state["status"] == "running":
            return "正在共享素材，等它跑完再操作"
    return None


def _preflight_drama(groups):
    """跑之前先检查每个计划名能不能匹配到短剧ID。

    匹配不到的计划一定会在「特定剧集」那步失败，与其跑到一半才炸，不如在上传时
    就告诉使用者是哪几个计划、名字对不上。
    """
    try:
        name_to_id, _ = load_series_map()
    except Exception as e:
        return [f"读不到「商品库-剧目」表: {e}"]

    # 注意：匹配不上时 resolve_series_from_campaign_name 是【抛异常】而不是返回空值，
    # 必须 try 住。第一版按「返回空值」写，预检自己先崩了。
    unmatched = []
    for (_advertiser_id, campaign_name), _rows in groups:
        try:
            _name, series_id = resolve_series_from_campaign_name(
                str(campaign_name), name_to_id
            )
        except Exception:
            series_id = None
        if not series_id:
            unmatched.append(str(campaign_name))
    if unmatched:
        return [
            f"这些计划名在「商品库-剧目」表里匹配不到剧目，跑到「特定剧集」那步会失败："
            + "、".join(unmatched)
        ]
    return []


def _preflight_identity(records):
    """表格里要选身份，但这台电脑上没有身份对照表 —— 上传时就说，别等跑到广告层。

    身份对照表按设计不进安装包（里面是真实账号信息，.gitignore 里就有它），
    所以每台新电脑上它一开始都不存在。别人电脑上第一次跑，每条广告一进广告层级
    就报 [Errno 2] No such file or directory 然后秒退。
    现在缺表不再让计划挂掉（见 identity_lookup.load_identity_map），但身份确实
    选不上，所以还是要在跑之前讲清楚。
    """
    want = [r for r in records if str(r.get("Identity_ID") or "").strip()]
    if not want:
        return []
    if not identity_file_exists():
        return [
            f"表格里有 {len(want)} 行填了 Identity_ID，但这台电脑上没有身份对照表 "
            "Identity_id.xlsx —— 这些广告的身份会选不上（其它步骤照常）。"
            "在下面【上传身份对照表】补一份就行。这个表不在安装包里是故意的："
            "里面是真实账号信息，不能进公开仓库。"
        ]

    # 表在，但表里查不到的 ID 也要提前说 —— 跑完才看到一堆警告没意义。
    mapping = load_identity_map()
    if not mapping:
        return [
            "身份对照表 Identity_id.xlsx 读出来是空的（页名/列顺序对不上？）。"
            "格式参考 examples/sample_identity_id.xlsx：第一列是显示名，第二列是 Identity_ID。"
        ]
    missing = sorted({
        str(r["Identity_ID"]).strip() for r in want
        if str(r["Identity_ID"]).strip() not in mapping
    })
    if missing:
        return [
            "这些 Identity_ID 在身份对照表里找不到对应名字，这几行的身份会选不上："
            + "、".join(missing[:10])
            + ("…" if len(missing) > 10 else "")
        ]
    return []


def _log_result(mode, publish, campaign_name, result):
    """把每个计划的结果追加到 logs/web_build.txt。

    网页版原来什么都不落盘，报错只在启动它的终端里滚过去——排查时只能靠使用者
    截图，截图太大还传不过去。命令行入口一直有 drama_build.txt，网页版也该有。
    """
    from src.config import LOGS_DIR

    try:
        LOGS_DIR.mkdir(parents=True, exist_ok=True)
        with open(str(LOGS_DIR / "web_build.txt"), "a", encoding="utf-8") as f:
            f.write(f"\n=== {campaign_name} ===\n")
            f.write(f"模式={mode} 发布={'是' if publish else '否'}\n")
            f.write("结果=" + ("成功" if result.get("success") else "失败") + "\n")
            if result.get("error"):
                f.write(f"失败原因: {result['error']}\n")
            for w in result.get("warnings") or []:
                f.write(f"  ! {w}\n")
    except Exception:
        pass


def _advertiser_ids(groups):
    """表格里出现过的所有广告主 ID，去重、保持出现顺序。

    原来只拿第一行那个去检查（ensure_chinese_ui(records[0]["Advertiser ID"])），
    因为作者自己的表里从头到尾就一个广告主。别人的表里完全可能一次跑好几个广告主，
    那样第 2 个广告主没权限就要跑到一半才炸，前面已经真发布出去的钱收不回来。
    """
    seen, out = set(), []
    for (advertiser_id, _campaign_name), _rows in groups:
        aid = str(advertiser_id).strip()
        if aid and aid not in seen:
            seen.add(aid)
            out.append(aid)
    return out


def _preflight_access(page, groups, mode):
    """开跑之前，把表格里每一个广告主 ID 都打开确认一遍。

    这一步是给别人用之后加的。以前的失败长这样：跑起来 -> 卡在语言检测 45 秒 ->
    报一句「页面可能没加载出来，或者当前登录账号没有这个广告主的权限」，
    使用者只看到「无权限操作」四个字，不知道是该登录、该换账号还是该改表格。
    现在在动手之前就把每个 ID 分别验一遍，报错直接说清是哪个 ID、哪种原因。
    """
    label = MODES[mode]["label"]
    ids = _advertiser_ids(groups)
    print(f"[预检] 表格里共 {len(ids)} 个广告主 ID: {', '.join(ids)}", flush=True)
    for aid in ids:
        access = check_advertiser_access(page, aid)
        if access.get("state") != "ok":
            raise PermissionError(describe_access(access, aid, label))
        print(f"[预检] 广告主 {aid} ✓ 可以打开（界面语言 {access.get('lang')}）",
              flush=True)


def _friendly_fatal(exc):
    """把致命异常翻成使用者看得懂的话；看不懂的才附完整堆栈。

    浏览器被关掉和没权限是最常见的两种，它们都不是「程序坏了」，
    甩一整个 Playwright 堆栈只会让人以为要改代码。
    """
    if is_browser_closed_error(exc):
        return ("浏览器窗口被关掉了，这次没跑完。\n"
                "跑的过程中请不要关那个自动打开的窗口——它就是干活的地方。\n"
                "已经发布出去的计划不会回滚，重跑之前请先去后台确认一下建到哪了。")
    if isinstance(exc, PermissionError):
        return str(exc)
    return traceback.format_exc()


def _run_build(xlsx_path, publish, mode, unique_creatives=False):
    try:
        records = load_rows(xlsx_path, mode=mode)
        groups = group_by_campaign(records)
        with state_lock:
            run_state["status"] = "running"
            run_state["mode"] = mode
            run_state["total"] = len(groups)
            run_state["completed"] = 0
            run_state["results"] = []
            run_state["current"] = None
            run_state["fatal_error"] = None

        profile = MODES[mode]["profile"]
        series_map = None
        if mode in ("drama", "episode"):
            try:
                series_map, _ = load_series_map()
            except Exception:
                # 端计划没有剧目对照表也能跑（退回按 '-' 拆计划名首段并警告）；
                # 商品库那条流程是硬依赖，让它照旧抛出去
                if mode == "drama":
                    raise

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(profile),
                headless=False,
                locale=LOCALE,
                extra_http_headers={"Accept-Language": ACCEPT_LANGUAGE},
                viewport={"width": 1600, "height": 1000},
            )
            page = context.pages[0] if context.pages else context.new_page()
            creative_usage = {}

            # 先确认登录 + 每个广告主 ID 的权限，再谈别的
            _preflight_access(page, groups, mode)

            if mode in ("drama", "episode") and records:
                # 整套定位都依赖中文文案，界面变英文时每一步都会失败，所以先拦住。
                # 实测界面会在中英文之间自己来回切，每次跑都要确认。
                from src.drama.pages.campaign_page import ensure_chinese_ui

                ensure_chinese_ui(page, str(records[0]["Advertiser ID"]).strip())

            for (advertiser_id, campaign_name), rows in groups:
                with state_lock:
                    run_state["current"] = campaign_name

                budget = rows[0]["Budget"]
                if mode == "episode":
                    result = build_episode_campaign(
                        page,
                        str(advertiser_id),
                        str(campaign_name),
                        budget,
                        rows,
                        publish=publish,
                        creative_usage=creative_usage,
                        series_name_map=series_map,
                    )
                elif mode == "drama":
                    result = build_drama_campaign(
                        page,
                        str(advertiser_id),
                        str(campaign_name),
                        budget,
                        rows,
                        publish=publish,
                        creative_usage=creative_usage,
                        series_map=series_map,
                    )
                else:
                    result = build_campaign_group(
                        page,
                        str(advertiser_id),
                        str(campaign_name),
                        budget,
                        rows,
                        publish=publish,
                        creative_usage=creative_usage,
                        unique_creatives=unique_creatives,
                    )
                result["campaign_name"] = campaign_name
                result["ad_group_count"] = len(rows)
                # 这条计划来自表格的第几行（表头算第 1 行）。界面上标在计划名前面，
                # 报错时能直接回表格定位。一个计划占多行时全部列出。
                result["excel_rows"] = [
                    r.get("_excel_row") for r in rows if r.get("_excel_row")
                ]
                _log_result(mode, publish, campaign_name, result)

                with state_lock:
                    run_state["results"].append(result)
                    run_state["completed"] += 1

            context.close()

        with state_lock:
            run_state["status"] = "done"
            run_state["current"] = None

    except Exception as e:
        message = _friendly_fatal(e)
        with state_lock:
            run_state["status"] = "error"
            run_state["fatal_error"] = message
        try:
            from src.config import LOGS_DIR

            LOGS_DIR.mkdir(parents=True, exist_ok=True)
            with open(str(LOGS_DIR / "web_build.txt"), "a", encoding="utf-8") as f:
                f.write(f"\n=== 整体停了（mode={mode}）===\n{message}\n")
                # 日志里永远留一份完整堆栈，界面上只给人话
                f.write(f"\n--- 完整堆栈 ---\n{traceback.format_exc()}\n")
        except Exception:
            pass


def _login_worker(mode):
    """开一个浏览器窗口让使用者登录自己的 BC 账号，登录态存进这个模式的 profile。

    和搭建用的是同一个 profile 目录，所以必须互斥（见 _busy_reason）。
    """
    profile = MODES[mode]["profile"]
    label = MODES[mode]["label"]
    try:
        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(profile),
                headless=False,
                locale=LOCALE,
                extra_http_headers={"Accept-Language": ACCEPT_LANGUAGE},
                viewport={"width": 1600, "height": 1000},
            )
            page = context.pages[0] if context.pages else context.new_page()

            def progress(waited, url):
                with state_lock:
                    login_state["message"] = (
                        f"已等待 {waited} 秒，请在打开的窗口里登录（当前 {url[:60]}）"
                    )

            found = wait_for_login(page, timeout_seconds=600, on_progress=progress)
            context.close()

        with state_lock:
            if found:
                login_state["status"] = "done"
                login_state["message"] = (
                    f"登录成功（检测到「{found}」）。{label}的登录态已保存，可以开始搭建了。"
                )
            else:
                login_state["status"] = "error"
                login_state["message"] = (
                    "等了 10 分钟还没检测到登录成功。如果你其实已经登进去了，"
                    "可以直接关掉这个窗口再点一次【检查登录状态】。"
                )
    except Exception as e:
        with state_lock:
            login_state["status"] = "error"
            if is_browser_closed_error(e):
                login_state["message"] = (
                    "登录窗口被关掉了。如果登录已经完成，登录态是保存下来的，"
                    "直接开始搭建就行；没登完的话再点一次【登录 / 换账号】。"
                )
            else:
                login_state["message"] = f"登录窗口出错: {e}"


@app.route("/account")
def account():
    """当前模式的登录状态（弱判断：只看 profile 里有没有登录过的痕迹）。"""
    mode = request.args.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400
    with state_lock:
        login = dict(login_state)
    return jsonify(
        {
            "ok": True,
            "mode": mode,
            "label": MODES[mode]["label"],
            "has_login": profile_has_login(MODES[mode]["profile"]),
            "login": login,
            "busy": _busy_reason(),
        }
    )


@app.route("/account/login", methods=["POST"])
def account_login():
    payload = request.json if request.is_json else {}
    mode = payload.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400

    busy = _busy_reason()
    if busy:
        return jsonify({"ok": False, "error": busy}), 400

    with state_lock:
        login_state.update(
            status="waiting",
            mode=mode,
            message="正在打开登录窗口……请在弹出的浏览器里登录你自己的 BC 账号。",
        )
    threading.Thread(target=_login_worker, args=(mode,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/account/logout", methods=["POST"])
def account_logout():
    """清掉当前模式的登录态 —— 换成另一个人/另一个 BC 账号时用。

    做法是把 profile 目录整个挪走（留一份上一次的），下次打开就是全新未登录状态。
    换账号必须走这一步：直接在浏览器里切账号，残留 cookie 经常会把人自动登回旧账号，
    看着像换了、跑起来还是旧的。
    """
    payload = request.json if request.is_json else {}
    mode = payload.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400

    busy = _busy_reason()
    if busy:
        return jsonify({"ok": False, "error": busy}), 400

    try:
        backup = clear_profile(MODES[mode]["profile"])
    except Exception as e:
        return jsonify({"ok": False, "error": f"清理登录态失败: {e}"}), 500

    with state_lock:
        login_state.update(
            status="idle",
            mode=mode,
            message=f"已退出登录（上一次的登录态备份在 {backup}）。点【登录 / 换账号】登新账号。",
        )
    return jsonify({"ok": True, "backup": backup})


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "没有选择文件"}), 400
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "请上传 .xlsx 文件"}), 400

    mode = request.form.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400

    f.save(str(UPLOAD_PATH))

    try:
        records = load_rows(str(UPLOAD_PATH), mode=mode)
        groups = group_by_campaign(records)
    except Exception as e:
        return jsonify({"ok": False, "error": f"表格格式有问题: {e}"}), 400

    warnings = _preflight_drama(groups) if mode == "drama" else []
    warnings += _preflight_identity(records)

    # 把表格里的广告主 ID 回显出来。给别人用之后这一项很重要：使用者能一眼看出
    # 程序读到的是不是他自己那个 ID，而不是跑起来才发现读错了列。
    advertiser_ids = _advertiser_ids(groups)
    if not profile_has_login(MODES[mode]["profile"]):
        warnings.append(
            f"浏览器里还没登录过（{MODES[mode]['label']}）。"
            "请先点上面的【登录 / 换账号】，用你自己的 BC 账号登录，再开始搭建。"
        )

    return jsonify(
        {
            "ok": True,
            "filename": f.filename,
            "row_count": len(records),
            "campaign_count": len(groups),
            "advertiser_ids": advertiser_ids,
            "warnings": warnings,
        }
    )


@app.route("/upload-identity", methods=["POST"])
def upload_identity():
    """上传身份对照表。

    为什么要有这个：这份表不进仓库也不进安装包（真实账号信息），所以每台新电脑上
    都得单独补一份。以前只能靠人工把文件拷到程序文件夹根目录——别人电脑上装完
    第一次跑，每条广告都在广告层级秒退，就是因为没人告诉他要拷这个文件。
    """
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "没有选择文件"}), 400
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "请上传 .xlsx 文件"}), 400

    # 先存到临时位置验一遍再就位，别用一个读不出来的文件把原来好的覆盖掉。
    #
    # 临时文件名【必须还是 .xlsx】：openpyxl 按扩展名判断格式，存成
    # Identity_id.xlsx.tmp 的话每一次上传都会被自己的校验拒掉
    #   openpyxl does not support .tmp file format
    # ——第一版就是这么写的，测试里三个用例全被拒才发现。
    tmp = IDENTITY_FILE.with_name("Identity_id.__uploading__.xlsx")
    f.save(str(tmp))
    try:
        import openpyxl

        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
        pairs = 0
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if row and len(row) >= 2 and row[0] is not None and row[1] is not None:
                pairs += 1
    except Exception as e:
        tmp.unlink(missing_ok=True)
        return jsonify({"ok": False, "error": f"这个 .xlsx 读不出来: {e}"}), 400

    if pairs == 0:
        tmp.unlink(missing_ok=True)
        return jsonify({
            "ok": False,
            "error": "表里没读到任何「显示名 + Identity_ID」。"
                     "格式参考 examples/sample_identity_id.xlsx："
                     "第一行是标题，第一列显示名，第二列 Identity_ID。",
        }), 400

    tmp.replace(IDENTITY_FILE)
    return jsonify({"ok": True, "count": pairs, "filename": f.filename})


@app.route("/identity-status")
def identity_status():
    if not identity_file_exists():
        return jsonify({"exists": False, "count": 0})
    return jsonify({"exists": True, "count": len(load_identity_map())})


@app.route("/run", methods=["POST"])
def run():
    with state_lock:
        if run_state["status"] == "running":
            return jsonify({"ok": False, "error": "已经在运行中，请等它跑完"}), 400

    if not UPLOAD_PATH.exists():
        return jsonify({"ok": False, "error": "还没上传表格"}), 400

    payload = request.json if request.is_json else {}
    publish = bool(payload.get("publish", True))
    mode = payload.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知模式: {mode}"}), 400

    # 「每组素材不同」只对小游戏有意义：短剧那条流程本来就是一个广告组一个广告、
    # 逐个挑素材，不存在「复制广告组把素材带走」这个问题。
    unique_creatives = bool(payload.get("unique_creatives", False)) and mode == "minigame"

    _reset_state()
    t = threading.Thread(
        target=_run_build,
        args=(str(UPLOAD_PATH), publish, mode, unique_creatives),
        daemon=True,
    )
    t.start()
    return jsonify({"ok": True})


def _run_share(mode, source_id, drama_names, account_names):
    """后台跑共享。和 _run_build 一个套路：开 profile、跑、把结果收进状态里。"""
    from src.share.config import PROFILES
    from src.share.runner import share_materials

    try:
        with state_lock:
            share_state.update(
                status="running", mode=mode, total=len(drama_names),
                completed=0, current=None, results=[], fatal_error=None,
            )

        with sync_playwright() as p:
            context = p.chromium.launch_persistent_context(
                user_data_dir=str(PROFILES[mode]),
                headless=False,
                locale=LOCALE,
                extra_http_headers={"Accept-Language": ACCEPT_LANGUAGE},
                viewport={"width": 1600, "height": 1000},
            )
            page = context.pages[0] if context.pages else context.new_page()

            # 源账号能不能打开，先验一遍 —— 打不开就没必要往下走
            access = check_advertiser_access(page, str(source_id).strip())
            if access.get("state") != "ok":
                raise ValueError(describe_access(access, source_id, MODES[mode]["label"]))

            def progress(i, total, drama):
                with state_lock:
                    share_state["current"] = drama
                    share_state["completed"] = i - 1

            results = share_materials(
                page, source_id, drama_names, account_names, on_progress=progress
            )
            context.close()

        with state_lock:
            share_state.update(status="done", results=results,
                               completed=len(results), current=None)
    except Exception as e:
        msg = _friendly_fatal(e)
        with state_lock:
            share_state.update(status="error", fatal_error=msg, current=None)


def _lines(text):
    """多行文本框 -> 去空行去空格的列表。"""
    return [ln.strip() for ln in str(text or "").splitlines() if ln.strip()]


@app.route("/share", methods=["POST"])
def share():
    busy = _busy_reason()
    if busy:
        return jsonify({"ok": False, "error": busy}), 400

    payload = request.json if request.is_json else {}
    mode = payload.get("mode", "minigame")
    if mode not in MODES:
        return jsonify({"ok": False, "error": f"未知登录态: {mode}"}), 400

    source_id = str(payload.get("source_id") or "").strip()
    dramas = _lines(payload.get("dramas"))
    accounts = _lines(payload.get("accounts"))

    missing = []
    if not source_id:
        missing.append("源账号 ID")
    if not dramas:
        missing.append("剧名（一行一个）")
    if not accounts:
        missing.append("目标账号名（一行一个）")
    if missing:
        return jsonify({"ok": False, "error": "还没填：" + "、".join(missing)}), 400

    t = threading.Thread(
        target=_run_share, args=(mode, source_id, dramas, accounts), daemon=True
    )
    t.start()
    return jsonify({"ok": True, "drama_count": len(dramas),
                    "account_count": len(accounts)})


@app.route("/share/status")
def share_status():
    with state_lock:
        return jsonify(dict(share_state))


@app.route("/status")
def status():
    with state_lock:
        return jsonify(dict(run_state))


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=PORT, debug=False)
