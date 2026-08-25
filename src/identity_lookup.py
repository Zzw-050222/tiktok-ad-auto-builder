import openpyxl

from src.config import PROJECT_ROOT

IDENTITY_FILE = PROJECT_ROOT / "Identity_id.xlsx"


def identity_file_exists():
    return IDENTITY_FILE.exists()


def load_identity_map():
    """身份表 → {Identity_ID: 显示名}。表不在、或格式不对，一律返回 {}，不抛异常。

    为什么刻意不抛：这个表【按设计】不进仓库也不进安装包——里面是真实账号信息，
    .gitignore 第 15 行就是它。所以每台新电脑上它一开始都不存在。
    别人电脑上第一次跑就撞上了：
        [Errno 2] No such file or directory: '.../Identity_id.xlsx'
    每条广告一进广告层级就秒退。

    而 builder 那边本来就把「选身份」当成非关键项（select_identity 失败只记一句
    警告，见 fill_ad_identity_copy_url 里那段说明）——只是「文件不存在」这一路
    发生在 try 之外，绕过了那道保护。让它返回空表，保护就重新生效了。
    """
    if not IDENTITY_FILE.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(IDENTITY_FILE, data_only=True)
    except Exception:
        # 文件在但打不开（下载坏了、其实是 .xls、正被 Excel 独占）。
        # 一样不该让整条计划挂掉。
        return {}

    # 不强求页名叫 Sheet1。README 让每个人按自己的账号【自己建】这份表，
    # 用 WPS / 中文版 Excel 新建，默认页名是「工作表1」「Sheet」之类，
    # 原来写死 wb["Sheet1"] 会直接 KeyError。
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception:
        return {}

    mapping = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        # 容忍空行和多余的列。原来写的是 `for name, identity_id in rows[1:]`，
        # 表里多一列就 ValueError——自己建表时非常容易多敲一列。
        if not row or len(row) < 2:
            continue
        name, identity_id = row[0], row[1]
        if identity_id is None or name is None:
            continue
        mapping[str(identity_id).strip()] = name
    return mapping


def resolve_identity(identity_id: str):
    mapping = load_identity_map()
    return mapping.get(identity_id.strip())
