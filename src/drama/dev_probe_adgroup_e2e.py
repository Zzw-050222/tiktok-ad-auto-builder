"""短剧广告组层端到端测试：从表格一行数据，走完整个广告组。

流程：
    确保中文 -> 选「TikTok 即时增长」-> 打开商品库开关 -> 填计划名和预算 -> 继续
    -> 广告组名 -> 选商品库（唯一可见）-> 特定剧集（切ID维度、搜、点圆圈、验证、添加）
    -> 选 TikTok Mini -> 目标 ROAS -> 地域 -> 继续 -> 广告层选创意素材

短剧从计划名里用最长前缀匹配推出（见 src/drama/series_lookup.py）。
地域和 ROAS 直接复用小游戏那套，操作一模一样。

**不发布**，走到「继续」为止，结束退出草稿。

用法：
    venv/bin/python3 -m src.drama.dev_probe_adgroup_e2e [表格路径]
"""

import sys
import traceback

import openpyxl
from playwright.sync_api import sync_playwright

from src.config import ACCEPT_LANGUAGE, LOCALE, LOGS_DIR
from src.drama.config import DRAMA_BROWSER_PROFILE_DIR
from src.drama.pages.adgroup_page import (
    add_specific_episode,
    fill_ad_group_name,
    select_product_catalog,
    select_target_roas_drama,
    select_tiktok_mini,
)
from src.drama.pages.ad_page import select_drama_creatives
from src.drama.pages.campaign_page import enable_catalog_campaign, ensure_chinese_ui
from src.drama.series_lookup import load_series_map, resolve_series_from_campaign_name
from src.pages.adgroup_page import set_regions
from src.pages.campaign_page import (
    continue_step,
    fill_campaign_details,
    select_native_growth_objective,
    start_new_campaign,
)
from src.pages.common import exit_draft, wait_until
from src.region_lookup import resolve_regions

REPORT = []


def L(s=""):
    REPORT.append(s)
    print(s, flush=True)


def shot(page, name):
    try:
        page.screenshot(path=str(LOGS_DIR / f"drama_e2e_{name}.png"))
    except Exception:
        pass


def load_first_row(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    for r in rows[1:]:
        if r and not all(v is None for v in r):
            return dict(zip(hdr, r))
    raise ValueError("表里没有有效数据行")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "短剧-excel.xlsx"
    # 可选：第二个参数覆盖素材搜索词。正式跑用剧名，但有些剧还没上素材，
    # 测试时可以指定一个库里确实有素材的剧名来验证选素材这段流程。
    search_override = sys.argv[2].strip() if len(sys.argv) > 2 else None
    d = load_first_row(path)

    advertiser_id = str(d["Advertiser ID"]).strip()
    campaign_name = str(d["Campaign Name"]).strip()
    ad_group_name = str(d["Ad Group Name"]).strip()
    budget = d["Budget"]
    roas = d["roas_bid"]
    region_raw = str(d["Region"]).strip()
    tt_mini_id = str(d.get("TT Mini ID") or "").strip()
    mini_name = str(d.get("Mini Game Name") or "").strip()
    creative_count = int(d.get("Creative Number") or 1)

    name_to_id, _ = load_series_map()
    series_name, series_id = resolve_series_from_campaign_name(campaign_name, name_to_id)

    L(f"表格: {path}")
    L(f"  广告主   = {advertiser_id}")
    L(f"  计划名   = {campaign_name}")
    L(f"  广告组名 = {ad_group_name[:60]}")
    L(f"  短剧     = {series_name!r} -> {series_id}   （从计划名最长前缀匹配得出）")
    L(f"  预算={budget}  ROAS={roas}  地域={region_raw}  素材数={creative_count}")
    L("  publish=False，只搭草稿不发布\n")

    region_pairs, missing = resolve_regions(region_raw)
    L(f"  地域解析: {region_pairs}" + (f"  对照表里找不到: {missing}" if missing else ""))

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(DRAMA_BROWSER_PROFILE_DIR),
            headless=False, locale=LOCALE,
            extra_http_headers={"Accept-Language": ACCEPT_LANGUAGE},
            viewport={"width": 1600, "height": 1000},
        )
        page = context.pages[0] if context.pages else context.new_page()
        step = 0

        def mark(name):
            nonlocal step
            step += 1
            L(f"  步骤 {step:>2}  {name}")

        try:
            mark("确保中文界面")
            ensure_chinese_ui(page, advertiser_id)

            mark("打开新建流程")
            start_new_campaign(page, advertiser_id)

            mark("选「TikTok 即时增长」")
            select_native_growth_objective(page)

            mark("打开「设置商品库推广系列」开关")
            changed = enable_catalog_campaign(page)
            L(f"          -> {'刚打开' if changed else '本来就是开的'}")

            mark("填计划名和预算")
            at_campaign = fill_campaign_details(page, campaign_name, budget)
            L(f"          -> 预算在计划层={at_campaign}")

            mark("继续 -> 广告组页")
            continue_step(page)
            ok = wait_until(page,
                            lambda: page.get_by_text("广告组名称", exact=True).count() > 0,
                            timeout_seconds=90)
            L(f"          -> 进入广告组页={bool(ok)}")
            page.wait_for_timeout(1500)
            shot(page, "01_adgroup")

            mark("填广告组名称")
            fill_ad_group_name(page, ad_group_name)

            mark("选商品库（唯一可见的）")
            select_product_catalog(page, catalog_id=None)
            shot(page, "02_catalog")

            mark(f"特定剧集：切ID维度、搜 {series_id}、点圆圈、验证、添加")
            add_specific_episode(page, series_id=series_id, series_name=series_name)
            shot(page, "03_episode")

            mark(f"选 TikTok Mini（{mini_name} / {tt_mini_id}）")
            select_tiktok_mini(page, tt_mini_id=tt_mini_id, mini_name=mini_name)
            from src.drama.pages.adgroup_page import _mini_is_selected
            L(f"          -> Mini 选中确认: {_mini_is_selected(page, mini_name, tt_mini_id)}")
            # 同时看看「优化目标」和「选择价值类型」有没有跟着联动
            state = page.evaluate("""() => {
              const deep = (n) => { let s=''; const w=(x)=>{ if(!x)return;
                if(x.nodeType===3){s+=x.textContent;return;}
                if(x.shadowRoot)w(x.shadowRoot);
                for(const c of x.childNodes||[])w(c); }; w(n);
                return s.replace(/\\s+/g,' ').trim(); };
              const sv = (title) => {
                for (const sec of document.querySelectorAll('[data-testid="lego-section-item"]')) {
                  const h = sec.querySelector('[data-testid="lego-section-item-header"]');
                  if (!h || (h.innerText||'').trim().split('\\n')[0].trim() !== title) continue;
                  const c = sec.querySelector('[data-testid="lego-section-item-content"]');
                  if (c) return deep(c).slice(0, 40);
                }
                return null; };
              return {goal: sv('优化目标'), vt: sv('选择价值类型')};
            }""")
            L(f"          -> 优化目标={state['goal']!r} 选择价值类型={state['vt']!r}")
            shot(page, "04_mini")

            mark(f"填目标 ROAS = {roas}")
            select_target_roas_drama(page, roas)

            mark(f"选地域 {region_pairs}")
            failed = set_regions(page, region_pairs)
            L(f"          -> 未选中的地区: {failed if failed else '无'}")
            shot(page, "05_regions")

            mark("继续 -> 广告层")
            continue_step(page)
            page.wait_for_timeout(3000)
            shot(page, "06_after_continue")

            search_term = search_override or series_name
            mark(f"选创意素材：搜 {search_term!r}，挑 {creative_count} 个"
                 + ("（命令行覆盖了搜索词）" if search_override else ""))
            used = set()
            picked, wrapped = select_drama_creatives(
                page, search_term, creative_count, used_ids=used)
            L(f"          -> 选中 {picked} 个"
              + ("（素材不够，绕回头复用过）" if wrapped else ""))
            shot(page, "07_creatives")

            L(f"\n{'=' * 60}\n✓ 广告组层全流程跑通\n{'=' * 60}")

        except Exception as e:
            L(f"\n✗ 第 {step} 步之后失败: {type(e).__name__}: {e}")
            L(traceback.format_exc())
            shot(page, "ERROR")
        finally:
            with open(str(LOGS_DIR / "drama_e2e.txt"), "w", encoding="utf-8") as f:
                f.write("\n".join(REPORT))
            print(f"\n报告已写入 {LOGS_DIR / 'drama_e2e.txt'}")
            try:
                exit_draft(page)
            except Exception:
                pass
            context.close()


if __name__ == "__main__":
    main()
