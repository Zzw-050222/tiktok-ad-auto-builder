"""短剧 剧名 <-> Series ID 对照，以及从计划名里认出是哪一部剧。

对照表是 商品库-剧目.xlsx（118 行，与商品库里「118 件商品」一致），两列：
    series_id    TIKTOKSERIES002
    series_name  The Tyrant of Silvermoon

为什么要有这个模块：投放表 Sheet1 里【没有】剧名或 series_id 这一列，剧名只出现在
Campaign Name 的开头，例如
    The Tyrant of Silvermoon-Yutong-US-We Shorts-IAA-1-00729-zzw2
"""

import openpyxl

from src.config import PROJECT_ROOT

SERIES_FILE = PROJECT_ROOT / "商品库-剧目.xlsx"


def load_series_map():
    """返回 (name -> id, id -> name) 两个字典。"""
    wb = openpyxl.load_workbook(SERIES_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    name_to_id, id_to_name = {}, {}
    for r in rows[1:]:
        if not r or r[0] is None or r[1] is None:
            continue
        sid, name = str(r[0]).strip(), str(r[1]).strip()
        if not sid or not name:
            continue
        name_to_id[name] = sid
        id_to_name[sid] = name
    return name_to_id, id_to_name


def resolve_series_from_campaign_name(campaign_name, name_to_id=None):
    """从计划名开头认出是哪一部短剧，返回 (series_name, series_id)。

    用【最长前缀匹配】而不是按 '-' 拆首段。按 '-' 拆看着简单，但实测对照表里有 4 部
    剧名本身带连字符，会被拆断：
        The Seventh-Year Intern        -> 拆成 'The Seventh'
        Married to My Ex-Fiance        -> 拆成 'Married to My Ex'
        The Ex-Wife Who Drank the Moon -> 拆成 'The Ex'
        The Stand-In Brid              -> 拆成 'The Stand'

    最长前缀匹配同时解决另一个问题：有 7 对剧名互为前缀/子串，例如
        Mark of the Moon  与  Mark of the Moon Season 2
        The Wolf Who Knelt 与 The Wolf Who Knelt (2)
    计划名 'Mark of the Moon Season 2-...' 两个都能匹配上，取【最长】的那个才对。

    认不出来就抛错，绝不返回一个「差不多」的结果——投错剧比不投更糟。
    """
    if name_to_id is None:
        name_to_id, _ = load_series_map()

    camp = str(campaign_name).strip()
    hits = [n for n in name_to_id if camp.startswith(n)]
    if not hits:
        raise ValueError(
            f"计划名 {camp!r} 的开头不匹配 {SERIES_FILE.name} 里的任何剧名。"
            "请确认计划名是以剧名开头的，或者在表里补上这一部。"
        )
    best = max(hits, key=len)
    return best, name_to_id[best]


def resolve_series_id(series_name, name_to_id=None):
    """剧名 -> series_id。要求完全相等。"""
    if name_to_id is None:
        name_to_id, _ = load_series_map()
    key = str(series_name).strip()
    if key not in name_to_id:
        raise ValueError(f"{SERIES_FILE.name} 里没有剧名 {key!r}")
    return name_to_id[key]
